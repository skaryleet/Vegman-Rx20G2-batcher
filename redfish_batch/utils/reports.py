from __future__ import annotations
from typing import List, Dict, Any
import os, json, csv
from .table_print import summarize_cpu_row, summarize_memory_row, summarize_raid_row, summarize_health_row

def _safe_filename(filename: str | None) -> str:
    filename = filename or ""
    return "".join(char if char.isalnum() or char in ("-", "_", ".") else "_" for char in filename) or "unknown"

def _collect_all_drives(result: Dict[str, Any]) -> list[dict]:
    drives = []
    raid = (result.get("summary") or {}).get("raid") if "summary" in result else result
    for system in (raid.get("systems") or []):
        for raid_controller in (system.get("raid") or []):
            controller = raid_controller.get("controller") or {}
            controller_id = controller.get("id") or controller.get("name")
            for drive in (raid_controller.get("drives") or []):
                drives.append({
                    "controller": controller_id,
                    "slot": drive.get("slot"),
                    "model": drive.get("model"),
                    "capacity_bytes": drive.get("capacity_bytes"),
                    "protocol": drive.get("protocol"),
                    "media_type": drive.get("media_type"),
                    "health": (drive.get("health") or None),
                })
    return drives
            
def collect_raid_disks_rows(results: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    successful_results = [result for result in results if result.get("ok")]
    for result in successful_results:
        host = result.get("host")
        ip_address = summarize_cpu_row(result).get("ip")  # безопасно достаём IP из base_url
        raid = (result.get("summary") or {}).get("raid") if "summary" in result else result
        for system in (raid.get("systems") or []):
            for raid_controller in (system.get("raid") or []):
                controller = raid_controller.get("controller") or {}
                controller_id = controller.get("id") or controller.get("name")
                for drive in (raid_controller.get("drives") or []):
                    rows.append({
                        "host": host,
                        "ip": ip_address,
                        "controller": controller_id,
                        "slot": drive.get("slot"),
                        "model": drive.get("model"),
                        "capacity_bytes": drive.get("capacity_bytes"),
                        "capacity_gib": round((drive.get("capacity_bytes") or 0) / (1024**3), 1) if drive.get("capacity_bytes") else None,
                        "protocol": drive.get("protocol"),
                        "media_type": drive.get("media_type"),
                        "health": (drive.get("health") or None),
                    })
    return rows

def save_raid_disks_csv(results: List[Dict[str, Any]], csv_path: str) -> None:
    rows = collect_raid_disks_rows(results)
    fieldnames = ["host", "ip", "controller", "slot", "model", "capacity_gib", "capacity_bytes", "protocol", "media_type", "health"]
    os.makedirs(os.path.dirname(csv_path) or ".", exist_ok=True)
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for row in rows:
            w.writerow({k: row.get(k) for k in fieldnames})

def save_raid_disks_xlsx(results: List[Dict[str, Any]], xlsx_path: str, sheet_name: str = "Disks") -> None:
    try:
        from openpyxl import Workbook
    except Exception as e:
        raise RuntimeError("Для экспорта в .xlsx нужен пакет 'openpyxl'. Установи: pip install openpyxl") from e

    rows = collect_raid_disks_rows(results)
    headers = ["host", "ip", "controller", "slot", "model", "capacity_gib", "capacity_bytes", "protocol", "media_type", "health"]

    os.makedirs(os.path.dirname(xlsx_path) or ".", exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]  # Excel ограничивает длину имени листа

    ws.append(headers)
    for r in rows:
        ws.append([r.get(h) for h in headers])

    # лёгкая авто-ширина
    for col_idx, hdr in enumerate(headers, start=1):
        max_len = max([len(str(hdr))] + [len(str(v)) for v in (r.get(hdr) for r in rows)])
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 50)

    wb.save(xlsx_path)

def save_all_reports(results: List[Dict[str, Any]], reports_dir: str = "reports") -> None:
    os.makedirs(reports_dir, exist_ok=True)
    ok = [x for x in results if x.get("ok")]
    for r in ok:
        cpu = summarize_cpu_row(r)
        mem = summarize_memory_row(r)
        raid = summarize_raid_row(r)
        health = summarize_health_row(r)
        fname = f"{_safe_filename(cpu.get('host'))}_{_safe_filename(cpu.get('ip'))}.json"
        path = os.path.join(reports_dir, fname)
        payload = {
            "host": cpu.get("host"),
            "ip": cpu.get("ip"),
            "base_url": r.get("base_url"),
            "model": cpu.get("model"),
            "serial_number": cpu.get("serial_number"),
            "bios_version": cpu.get("bios_version"),
            "cpu_model": cpu.get("cpu_model"),
            "total_cores": cpu.get("total_cores"),
            "ram_gib": mem.get("ram_gib"),
            "raid": {
                "controllers": raid.get("raid_controllers"),
                "volumes": raid.get("raid_volumes"),
                "drives": raid.get("raid_drives"),
                "present": bool(raid.get("raid_present")),
            },
            "health": {
                "system_health": health.get("system_health"),
                "system_state": health.get("system_state"),
                "power_state": health.get("power_state"),
            },
            "summary": r.get("summary"),
            "all_drives": _collect_all_drives(r)  # полный оригинал
        }
        with open(path, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)

def save_all_csv(results: List[Dict[str, Any]], csv_path: str) -> None:
    ok = [x for x in results if x.get("ok")]
    rows = []
    for r in ok:
        cpu = summarize_cpu_row(r)
        mem = summarize_memory_row(r)
        raid = summarize_raid_row(r)
        health = summarize_health_row(r)
        rows.append({
            "host": cpu.get("host"),
            "ip": cpu.get("ip"),
            "model": cpu.get("model"),
            "serial_number": cpu.get("serial_number"),
            "bios_version": cpu.get("bios_version"),
            "cpu_model": cpu.get("cpu_model"),
            "total_cores": cpu.get("total_cores"),
            "ram_gib": mem.get("ram_gib"),
            "raid_controllers": raid.get("raid_controllers"),
            "raid_volumes": raid.get("raid_volumes"),
            "raid_drives": raid.get("raid_drives"),
            "system_health": health.get("system_health"),
        })
    fieldnames = list(rows[0].keys()) if rows else [
        "host","ip","model","serial_number","bios_version","cpu_model","total_cores",
        "ram_gib","raid_controllers","raid_volumes","raid_drives","system_health"
    ]
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for row in rows:
            w.writerow(row)
            
def collect_accounts_rows(results: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    ok = [x for x in results if x.get("ok")]
    for r in ok:
        host = r.get("host")
        # IP можно достать через summarize_cpu_row, но здесь проще разобрать base_url:
        from urllib.parse import urlparse
        try:
            ip = (urlparse(r.get("base_url") or "").netloc or "").split(":")[0]
        except Exception:
            ip = None
        for a in (r.get("accounts") or []):
            rows.append({
                "host": host,
                "ip": ip,
                "account_uri": a.get("account_uri"),
                "id": a.get("id"),
                "user_name": a.get("user_name"),
                "role_id": a.get("role_id"),
                "enabled": a.get("enabled"),
                "locked": a.get("locked"),
            })
    return rows

def save_accounts_csv(results: List[Dict[str, Any]], csv_path: str) -> None:
    rows = collect_accounts_rows(results)
    os.makedirs(os.path.dirname(csv_path) or ".", exist_ok=True)
    fieldnames = ["host", "ip", "user_name", "role_id", "enabled", "locked", "id", "account_uri"]
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for row in rows:
            w.writerow({k: row.get(k) for k in fieldnames})

def save_accounts_xlsx(results: List[Dict[str, Any]], xlsx_path: str, sheet_name: str="Accounts") -> None:
    try:
        from openpyxl import Workbook
    except Exception as e:
        raise RuntimeError("Для экспорта в .xlsx нужен пакет 'openpyxl' (pip install openpyxl)") from e

    rows = collect_accounts_rows(results)
    headers = ["host", "ip", "user_name", "role_id", "enabled", "locked", "id", "account_uri"]

    os.makedirs(os.path.dirname(xlsx_path) or ".", exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h) for h in headers])
    # Авто-ширина столбцов
    for col_idx, hdr in enumerate(headers, start=1):
        max_len = max([len(str(hdr))] + [len(str(v)) for v in (r.get(hdr) for r in rows)])
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 50)
    wb.save(xlsx_path)
